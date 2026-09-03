#!/usr/bin/env python
"""
excelrouter · 「到人」二级拆分（v2.6.1，scripts 包装层，不改动 vendor/core）。

为什么独立成模块：vendor/core/splitter.py 的 --to-person 仅支持「单一全局人字段 + 目录输入」，
对「单文件多 sheet、各 sheet 人字段不同」的场景（如 结对子维表→1看管人，渠道维表→看管 渠道管理员）
无能为力，且 UPSTREAM.md 明确禁止手改 vendor。于是这里在 CLI 包装层补一道：
先让 vendor 把源文件按主字段（网格）拆好（格式已完整保留），再读这些网格文件，
按「每个 sheet 各自的人字段」把行拆给对应的人，写成 {网格}/到人/{姓名}_{网格}.xlsx。

格式保留策略：网格文件已由 vendor 复制好表头/数据样式，这里逐格搬运 value + 字体/填充/
边框/对齐/数字格式，并复制列宽与（所搬行范围内的）合并单元格，视觉上与源一致。
"""

import os
import shutil
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter

from core.utils import safe_filename, soft_clean

# 自动选人字段的关键词优先级（命中即采用，越靠前越优先）
_AUTO_KEYWORDS = ["接收人", "看管人", "管理员", "负责人", "对接人", "主管", "队长"]


def parse_person_spec(spec):
    """把 --to-person 的值解析成 {sheet_name: col} 或 {'*': col} 或 {'*': 'auto'}。

    - "姓名"                       -> {'*': '姓名'}            （所有 sheet 共用一个字段）
    - "auto"                       -> {'*': 'auto'}            （每个 sheet 自动挑人字段）
    - "结对子维表:1看管人,渠道维表:看管 渠道管理员"
                                   -> {'结对子维表':'1看管人', '渠道维表':'看管 渠道管理员'}
    """
    spec = (spec or "").strip()
    if not spec:
        return None
    if ":" not in spec:
        return {"*": spec}
    mapping = {}
    for part in spec.split(","):
        part = part.strip()
        if not part or ":" not in part:
            return None  # 混合语法视为非法，退回 None 让上层报错
        sn, col = part.split(":", 1)
        mapping[sn.strip()] = col.strip()
    return mapping


def _find_header_col(ws, target):
    """在 ws 前 20 行里找表头行与列下标。target 为列名或 'auto'。返回 (h_row, col_idx) 或 (None, None)。"""
    max_scan = min(ws.max_row, 20)
    for r in range(1, max_scan + 1):
        cells = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        nonempty = [c for c in cells if soft_clean(c) != ""]
        if not nonempty:
            continue
        # 候选表头行：找到 target 列
        names = [soft_clean(c) for c in cells]
        if target == "auto":
            # 收集所有 (关键词, 列号, 列名) 命中，再统一择优，避免：
            #   1) 命中辅助/备注列（如「如需要修改F列，请在此填写正确的看管人姓名」）
            #   2) 因关键词先后顺序，跳过后面更准确的列（如「管理员」才对应「看管 渠道管理员」）
            helper_tokens = ("修改", "填写", "备注", "说明", "注意", "请填", "此处")
            matches = []
            for kw in _AUTO_KEYWORDS:
                for i, nm in enumerate(names):
                    if nm and kw in nm:
                        matches.append((kw, i + 1, nm))
            if matches:
                filtered = [m for m in matches
                            if not any(t in m[2] for t in helper_tokens)]
                cand = filtered if filtered else matches

                def _score(m):
                    kw, idx, nm = m
                    exact = 0 if nm == kw else 1  # 列名==关键词 优先（精确匹配）
                    return (exact, len(nm), idx)   # 其次列名更短、列号更靠前
                cand.sort(key=_score)
                return r, cand[0][1]
        else:
            for i, nm in enumerate(names):
                if nm == target or (target and (target in nm or nm in target)):
                    return r, i + 1
    return None, None


def _resolve_col(sheet_name, mapping):
    """根据 mapping 解析某个 sheet 的人字段名（支持 sheet 专属 + '*' 兜底 + auto 透传）。"""
    if mapping is None:
        return None
    if sheet_name in mapping:
        return mapping[sheet_name]
    return mapping.get("*")


def _copy_row(src_ws, src_r, dst_ws, dst_r, max_col):
    """把 src_ws 第 src_r 行按格搬到 dst_ws 第 dst_r 行，连值带样式一起。"""
    for c in range(1, max_col + 1):
        s_cell = src_ws.cell(src_r, c)
        d_cell = dst_ws.cell(dst_r, c)
        d_cell.value = s_cell.value
        d_cell.font = copy(s_cell.font)
        d_cell.fill = copy(s_cell.fill)
        d_cell.border = copy(s_cell.border)
        d_cell.alignment = copy(s_cell.alignment)
        d_cell.number_format = s_cell.number_format


def _write_person_file(sheet_entries, out_path):
    """把 {sheet_name: [(源工作簿, 源行号), ...]} 写成一个人的 xlsx。

    同一个 sheet 名可能来自多个源工作簿（目录输入时，一个网格下会有多个源文件），
    这里按 sheet 名合并到同一张输出 sheet：表头取第一个来源的，数据行依次续写。
    """
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)
    used_titles = set()
    for sn, entries in sheet_entries.items():
        if not entries:
            continue
        first_wb = entries[0][0]
        src_ws = first_wb[sn]
        h_row = _locate_header(src_ws)
        if h_row is None:
            continue
        title, base, n = sn, sn, 1
        while title in used_titles:
            n += 1
            title = f"{base}({n})"
        used_titles.add(title)
        dst_ws = dst_wb.create_sheet(title=title)
        # 列宽取第一个来源的
        for c in range(1, src_ws.max_column + 1):
            w = src_ws.column_dimensions[get_column_letter(c)].width
            if w is not None:
                dst_ws.column_dimensions[get_column_letter(c)].width = w
        # 表头
        _copy_row(src_ws, h_row, dst_ws, 1, src_ws.max_column)
        # 数据行（可能跨多个源工作簿）
        dst_r = 2
        for wb, src_r in entries:
            ws = wb[sn]
            _copy_row(ws, src_r, dst_ws, dst_r, ws.max_column)
            dst_r += 1
        # 合并单元格只搬表头行内的那部分：数据区合并单元格本来就不承诺保留
        # （见 README「诚实边界」），而且跨来源后源行号已不等于目标行号，
        # 按源坐标 merge 会串到错误的行上——旧实现这里是有 bug 的。
        for mc in src_ws.merged_cells.ranges:
            if mc.min_row >= h_row and mc.max_row <= h_row:
                dst_ws.merge_cells(
                    start_row=mc.min_row - h_row + 1, start_column=mc.min_col,
                    end_row=mc.max_row - h_row + 1, end_column=mc.max_col,
                )
    dst_wb.save(out_path)


def _locate_header(ws):
    """返回第一个非空表头行号（1 基），找不到返回 None。"""
    max_scan = min(ws.max_row, 20)
    for r in range(1, max_scan + 1):
        if any(soft_clean(ws.cell(r, c).value) != "" for c in range(1, ws.max_column + 1)):
            return r
    return None


def _iter_grid_files(grid_dir):
    """遍历产出目录，找出每个「网格工作簿」及其「到人」归属目录。

    内核在不同输入模式下布局不一样，只 listdir 顶层会漏掉目录输入的结果：
      单文件输入            → {out}/{网格}.xlsx
      目录输入 merge=False  → {out}/{网格}/汇总/{网格}_{源文件名}.xlsx
    两种情况统一归到 {网格}/到人/ 下。
    """
    for dirpath, dirnames, filenames in os.walk(grid_dir):
        # 不进入上一轮已经产出的「到人」目录，避免自我递归/重复拆分
        dirnames[:] = [d for d in dirnames if d != "到人"]
        for fn in sorted(filenames):
            low = fn.lower()
            if not low.endswith(".xlsx") or fn.startswith("~$") or low.endswith("__tmp__.xlsx"):
                continue
            full = os.path.join(dirpath, fn)
            parent = os.path.dirname(full)
            if os.path.basename(parent) == "汇总":
                grid_root = os.path.dirname(parent)          # {out}/{网格}
            else:
                grid_root = os.path.join(parent, os.path.splitext(fn)[0])  # {out}/{网格}
            yield full, grid_root


def person_split(grid_dir, person_map, log_fn=None):
    """对 vendor 产出的结果做「到人」拆分：每个网格目录下生成 {网格}/到人/{人}_{网格}.xlsx。

    同一网格下的多个源文件（目录输入时很常见）会合并到同一个人的文件里，
    而不是一个源文件一份。返回写入的「到人」文件数。
    """
    if not person_map:
        return 0

    # 先按网格归堆，跨源文件合并才有依据
    groups = {}
    for path, grid_root in _iter_grid_files(grid_dir):
        groups.setdefault(grid_root, []).append(path)

    count = 0
    for grid_root, paths in groups.items():
        grid_name = os.path.basename(grid_root.rstrip(os.sep)) or "结果"
        buckets = {}          # person -> {sheet_name: [(wb, src_row), ...]}
        opened = []
        try:
            for path in paths:
                try:
                    wb = openpyxl.load_workbook(path, data_only=True)
                except Exception as e:
                    if log_fn:
                        log_fn(f"  ⚠️ 读取网格文件「{os.path.basename(path)}」失败：{e}")
                    continue
                opened.append(wb)
                for sn in wb.sheetnames:
                    col = _resolve_col(sn, person_map)
                    if not col:
                        continue
                    h_row, c_idx = _find_header_col(wb[sn], col)
                    if c_idx is None:
                        if log_fn and col != "auto":
                            log_fn(f"  ⚠️ 网格「{grid_name}」sheet「{sn}」里没找到人字段"
                                   f"「{col}」，跳过该 sheet")
                        continue
                    ws = wb[sn]
                    for r in range(h_row + 1, ws.max_row + 1):
                        v = ws.cell(r, c_idx).value
                        if v is None or soft_clean(v) == "":
                            continue
                        buckets.setdefault(str(v), {}).setdefault(sn, []).append((wb, r))
            for person, sheets in buckets.items():
                out_path = os.path.join(grid_root, "到人",
                                        f"{safe_filename(person)}_{grid_name}.xlsx")
                try:
                    _write_person_file(sheets, out_path)
                    count += 1
                except Exception as e:
                    if log_fn:
                        log_fn(f"  ⚠️ 写「{out_path}」失败：{e}")
        finally:
            for wb in opened:
                try:
                    wb.close()
                except Exception:
                    pass

    if log_fn and count:
        log_fn(f"  👤 到人拆分新增 {count} 个文件")
    if log_fn and not count and person_map:
        # 一声不吭地交白卷最伤信任：宁可吵一点，也要让人知道没拆出东西
        log_fn("  ⚠️ 到人拆分没有产出任何文件——请核对人字段名是否正确"
               "（可用 --to-person auto 让工具自动挑，或先用 er_inspect.py 看有哪些列）")
    return count


def flatten_output(output_path, output_root, log_fn=None):
    """--no-timestamp：把带时间戳的子目录内容提升到 output_root，返回 output_root。"""
    if not output_path or os.path.abspath(output_path) == os.path.abspath(output_root):
        return output_path
    output_root = os.path.abspath(output_root)
    output_path = os.path.abspath(output_path)
    os.makedirs(output_root, exist_ok=True)
    try:
        for item in os.listdir(output_path):
            src = os.path.join(output_path, item)
            dst = os.path.join(output_root, item)
            if os.path.exists(dst):
                if os.path.isdir(dst) and os.path.isdir(src):
                    for sub in os.listdir(src):
                        shutil.move(os.path.join(src, sub), os.path.join(dst, sub))
                    continue
                else:
                    # 同名冲突：覆盖（先删后移）
                    if os.path.isdir(dst):
                        shutil.rmtree(dst)
                    else:
                        os.remove(dst)
            shutil.move(src, dst)
    except Exception as e:
        if log_fn:
            log_fn(f"  ⚠️ 扁平化输出目录失败：{e}")
    # 清理空的时间戳目录（直接重试 rmdir，规避 Windows 刚关闭/移动文件时的瞬时目录锁）
    import time as _t
    removed = False
    for _ in range(10):
        try:
            os.rmdir(output_path)
            removed = True
            break
        except OSError:
            try:
                if os.listdir(output_path):
                    break  # 目录非空，停止（不误删残留文件）
            except OSError:
                pass
            _t.sleep(0.3)
    # 删不掉不致命（内容已经就位），但必须说一声：否则用户会同时看到一个空的时间戳
    # 目录和已平铺的结果，不知道该看哪边。常见于文件刚被移动、目录仍被占用，
    # 或删除动作被杀软/沙箱拦截。
    if not removed and log_fn:
        log_fn(f"  ⚠️ 结果已平铺到 {output_root}，但原时间戳目录没能自动删除："
               f"{output_path}（不影响使用，手动删掉即可）")
    return output_root
