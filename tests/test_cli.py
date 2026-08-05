"""三个 er_*.py CLI 的冒烟测试：以子进程方式跑（就是 Claude 实际会怎么调），
解析 stdout 最后一行 JSON。样本构造沿用 excel-router 上游 tests/ 的写法
（表头故意不放第一行、带样式、含合计行），确保 CLI 包装层没有偷懒漏传参数。
"""
import json
import os
import subprocess
import sys

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPTS = os.path.join(REPO_ROOT, "skills", "excelrouter", "scripts")

HEADERS = ["工号", "姓名", "部门", "城市", "金额"]


def _make_book(path, rows, title="月度报表"):
    """表头在第 2 行、第 1 行是大标题——专测自动表头识别不是碰巧读对了第一行。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = title
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for i, r in enumerate(rows, start=3):
        for c, v in enumerate(r, 1):
            ws.cell(row=i, column=c, value=v)
    wb.save(path)


def _run(script, *args):
    """跑一个 er_*.py，返回 (returncode, json_obj, stderr_text)。"""
    proc = subprocess.run(
        [sys.executable, os.path.join(SCRIPTS, script), *args],
        capture_output=True, text=True, encoding="utf-8",
    )
    stdout_lines = [line for line in proc.stdout.splitlines() if line.strip()]
    assert stdout_lines, f"stdout 应至少有一行 JSON，实际为空。stderr:\n{proc.stderr}"
    obj = json.loads(stdout_lines[-1])
    return proc.returncode, obj, proc.stderr


@pytest.fixture
def two_books(tmp_path):
    inp = tmp_path / "in"
    out = tmp_path / "out"
    inp.mkdir()
    out.mkdir()
    _make_book(inp / "A.xlsx", [
        ["001", "张三", "销售部", "北京", 100],
        ["002", "李四", "销售部", "北京", 200],
        ["003", "王五", "销售部", "上海", 150],
        ["004", "赵六", "技术部", "北京", 300],
        ["", "合计", "", "", 750],
    ])
    _make_book(inp / "B.xlsx", [
        ["005", "钱七", "销售部", "广州", 120],
        ["006", "孙八", "技术部", "深圳", 220],
    ])
    return str(inp), str(out)


# ---------- er_inspect.py ----------

def test_inspect_single_file_detects_header_and_columns(tmp_path):
    f = tmp_path / "a.xlsx"
    _make_book(f, [["001", "张三", "销售部", "北京", 100]])
    rc, obj, _ = _run("er_inspect.py", "--input", str(f))
    assert rc == 0
    assert obj["ok"] is True
    assert obj["is_dir"] is False
    assert obj["excel_count"] == 1
    assert obj["header_row"] == 2
    assert obj["columns"] == HEADERS


def test_inspect_lists_values_and_skips_total_row(tmp_path):
    f = tmp_path / "a.xlsx"
    _make_book(f, [
        ["001", "张三", "销售部", "北京", 100],
        ["002", "李四", "技术部", "北京", 200],
        ["", "合计", "", "", 300],
    ])
    rc, obj, _ = _run("er_inspect.py", "--input", str(f), "--column", "部门")
    assert rc == 0
    assert obj["values"] == ["技术部", "销售部"]


def test_inspect_directory_counts_files(two_books):
    inp, _out = two_books
    rc, obj, _ = _run("er_inspect.py", "--input", inp)
    assert rc == 0
    assert obj["is_dir"] is True
    assert obj["excel_count"] == 2


def test_inspect_missing_input_reports_friendly_error(tmp_path):
    rc, obj, _ = _run("er_inspect.py", "--input", str(tmp_path / "not-exist.xlsx"))
    assert rc == 1
    assert obj["ok"] is False
    assert "找不到" in obj["error"]


# ---------- er_split.py ----------

def test_split_missing_by_reports_friendly_error(two_books):
    inp, out = two_books
    rc, obj, _ = _run("er_split.py", "--input", inp, "--output", out)
    assert rc == 1
    assert obj["ok"] is False
    assert "--by" in obj["error"]


def test_split_dry_run_does_not_write_files(two_books):
    inp, out = two_books
    rc, obj, _ = _run("er_split.py", "--input", inp, "--output", out, "--by", "部门", "--dry-run")
    assert rc == 0
    assert obj["dry_run"] is True
    assert obj["config"]["split_column"] == "部门"
    assert os.listdir(out) == []


def test_split_produces_expected_output_files(two_books):
    inp, out = two_books
    rc, obj, stderr = _run("er_split.py", "--input", inp, "--output", out, "--by", "部门")
    assert rc == 0, stderr
    assert obj["ok"] is True
    output_path = obj["output_path"]
    assert os.path.isdir(output_path)
    # 默认 merge_across_files=False：按原表各自拆分，两个源文件各自产出「部门」子目录
    produced = set(os.listdir(output_path))
    assert "运行日志.txt" in produced
    # A.xlsx 和 B.xlsx 都含「销售部」，都产出该主取值目录
    assert "销售部" in produced or "销售部.zip" in produced


def test_split_merge_flag_merges_across_files(two_books):
    inp, out = two_books
    rc, obj, stderr = _run("er_split.py", "--input", inp, "--output", out, "--by", "部门", "--merge", "--no-zip")
    assert rc == 0, stderr
    output_path = obj["output_path"]
    # merge=True 时同取值跨文件合并为扁平文件：{部门}.xlsx 直接在输出根
    assert os.path.exists(os.path.join(output_path, "销售部.xlsx"))


def test_split_selected_values_filters(two_books):
    inp, out = two_books
    rc, obj, stderr = _run("er_split.py", "--input", inp, "--output", out, "--by", "部门",
                            "--values", "技术部", "--merge", "--no-zip")
    assert rc == 0, stderr
    output_path = obj["output_path"]
    assert os.path.exists(os.path.join(output_path, "技术部.xlsx"))
    assert not os.path.exists(os.path.join(output_path, "销售部.xlsx"))


# ---------- er_pdf_dist.py ----------

def _make_pdf(path, pages=1):
    from fpdf import FPDF
    pdf = FPDF()
    pdf.set_auto_page_break(False)
    for i in range(pages):
        pdf.add_page()
        pdf.set_font("Helvetica", size=14)
        pdf.text(30, 40, f"SOURCE page{i + 1}")
    pdf.output(str(path))


def _make_mapping(path, rows, headers=("GridName", "Pwd", "Receiver")):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def test_pdf_dist_list_columns(tmp_path):
    mp = tmp_path / "map.xlsx"
    _make_mapping(mp, [("GridA", "001234", "Zhang")])
    rc, obj, _ = _run("er_pdf_dist.py", "--mapping", str(mp), "--list-columns")
    assert rc == 0
    assert obj["columns"] == ["GridName", "Pwd", "Receiver"]


def test_pdf_dist_encrypts_and_writes_manifest(tmp_path):
    pdf_path = tmp_path / "report.pdf"
    _make_pdf(pdf_path)
    mp = tmp_path / "map.xlsx"
    _make_mapping(mp, [("GridA", "001234", "Zhang"), ("GridB", 8888, "Li")])
    out = tmp_path / "dist"
    rc, obj, stderr = _run(
        "er_pdf_dist.py", "--pdf", str(pdf_path), "--mapping", str(mp),
        "--grid-col", "GridName", "--password-col", "Pwd", "--receiver-col", "Receiver",
        "--output", str(out), "--no-watermark",
    )
    assert rc == 0, stderr
    assert obj["ok"] is True
    assert os.path.exists(obj["manifest"])
    assert os.path.exists(os.path.join(obj["output_path"], "GridA", "report.pdf"))

    from pypdf import PdfReader
    reader = PdfReader(os.path.join(obj["output_path"], "GridA", "report.pdf"))
    assert reader.is_encrypted


def test_pdf_dist_missing_columns_reports_friendly_error(tmp_path):
    pdf_path = tmp_path / "report.pdf"
    _make_pdf(pdf_path)
    mp = tmp_path / "map.xlsx"
    _make_mapping(mp, [("GridA", "001234", "Zhang")])
    rc, obj, _ = _run("er_pdf_dist.py", "--pdf", str(pdf_path), "--mapping", str(mp))
    assert rc == 1
    assert obj["ok"] is False
    assert "--grid-col" in obj["error"]
